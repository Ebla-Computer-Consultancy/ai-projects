from openpyxl import Workbook, load_workbook
#this script is to split all work sheets in microsoft excel sheet to separate excel files.
# Load the Excel workbook
workbook = load_workbook('Labfiles\data\mom\MM Knowledge Base(AutoRecovered).xlsx')

# Iterate over each worksheet in the workbook
for sheet_name in workbook.sheetnames:
    # Get the worksheet by name
    worksheet = workbook[sheet_name]
    
    # Create a new workbook and add the current worksheet to it
    new_workbook = Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = sheet_name
    
    # Iterate over the rows in the current worksheet and copy data
    for row in worksheet.iter_rows(values_only=True):
        new_worksheet.append(row)
    
    # Save the new workbook as a separate Excel file
    new_workbook.save(f'{sheet_name}.xlsx')